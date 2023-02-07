# %% [markdown]
# 

# %%
from os import listdir
from os.path import isfile, join
import pandas as pd

# %% [markdown]
# Read Exported files using pandas

# %%
import statistics

def getStatistics(data):
    meaN = "%.3f" % statistics.mean(data)
    stDev = "%.3f" % statistics.stdev(data)
    return meaN, stDev

# %%
def calcuateTPR(clear,total):
    tpr = clear/total * 100
    if 0 < tpr < 100:
        tpr = "%.2f" % tpr
    return tpr

# %%
def getTargetCqValues(lodFile):
    targets = dict()
    for i in lodFile.index:
        cqValue = lodFile.loc[i,"Cq"]
        sampleName = lodFile.loc[i,"Sample"]
        reporterDye = lodFile.loc[i,'Reporter']
        if pd.isna(sampleName):
            continue
        if sampleName == "PC" or sampleName == "NC":
            continue
        string = sampleName+"\t"+reporterDye
        targets.setdefault(string,[]).append(cqValue)
    return targets


# %%
def runMain(lodFile,mypath):
    targetThresholds={"VIC":38,"ABY":37,"FAM":38,"ALEXA 647":38,"JUN":33}
    targets = getTargetCqValues(lodFile=lodFile)

    
    df_list = pd.DataFrame()
    text = open(mypath+"/a.txt","a")

    for target in targets:
        s = target.split("\t")
        total = len(targets[target])
        clear = 0
        tpr = 0
        cqMaen = 0
        cqSd = 0
        data=[]
        for v in targets[target]:
            if v < 43:
                data.append(v)
            if v < targetThresholds[s[1]]:
                clear +=1
        tpr = calcuateTPR(clear=clear,total=total)
        
        if len(data) > 1:
            cqMaen , cqSd = getStatistics(data)
            
        text.write(s[0]+"\t"+s[1]+"\t"+str(clear)+"\t"+str(total)+"\t"+str(tpr)+"\t"+str(cqMaen)+"\t"+str(cqSd)+"\n")
        listL=[s[1],clear,total,tpr,cqMaen,cqSd]
        dfx = pd.DataFrame([listL],index=[s[0]],columns=['Target','Clean','Total','TPR','Mean','SD'])
        df_list = pd.concat([df_list,dfx])

    
    text.close()
    return df_list


# %%
from openpyxl import load_workbook

mypath = "D:/Covid_data_New/Yr2023/Jan/Jan31/"
writer = pd.ExcelWriter(mypath+"/aa.xlsx")
onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
for file in onlyfiles:
    if "xlsx" in file:
        pathOfFile = mypath+"/"+file
        success = True
        try:
            wb = load_workbook(pathOfFile, read_only=True)   # open an Excel file and return a workbook
        except:
            success=False
            print ("Skipping " + file)
        if "Primary_result" in wb.sheetnames and success:
            s = file.split("_")
            sheetName = s[4]+"_"+s[5]
            lodFile = pd.read_excel(pathOfFile,sheet_name="Primary_result",skiprows=26)
            lodFile = lodFile.replace(to_replace="UNDETERMINED",value=43)
            df = runMain(lodFile=lodFile,mypath=mypath)
            df.to_excel(writer,sheet_name = sheetName )
writer.close()


